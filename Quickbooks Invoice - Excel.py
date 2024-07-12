import re
import pdfplumber
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from tkinter import Tk, filedialog, messagebox

# Function to extract the invoice number from the PDF
def extract_invoice_number(pdf_path):
    reader = PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()

    # Use regex to find the invoice number
    match = re.search(r'Invoice #\s*(\d+)', text)
    if match:
        return match.group(1)
    else:
        return "Invoice number not found"

# Function to extract Bill To information from the PDF text
def extract_bill_to_info(pdf_text):
    bill_to_match = re.search(r'Bill To\s*(.+?)(?=\s{2,}|\n\s*\n|P\.O\. No\.|Terms|Rep|Project|Phone:)', pdf_text, re.DOTALL)
    if bill_to_match:
        bill_to_text = bill_to_match.group(1).strip()
        bill_to_lines = bill_to_text.split('\n')
        
        email = ''
        phone = ''
        client_name = ''
        company_lines = []
        
        print(f"Extracted Bill To lines: {bill_to_lines}")  # Debug print

        # Scan from the last line of the "Bill To" information
        for i in range(len(bill_to_lines) - 1, -1, -1):
            line = bill_to_lines[i].strip()
            print(f"Processing line: {line}")  # Debug print
            if re.match(r'\S+@\S+\.\S+', line):  # Line is an email
                email = line
            elif re.match(r'\d{3}-\d{3}-\d{4}', line):  # Line is a phone number
                phone = line
            elif email or phone:
                client_name = line
                company_lines = bill_to_lines[:i]
                break
            else:
                client_name = line
                company_lines = bill_to_lines[:i]
                break
        
        company = ' '.join(company_lines).strip()
    else:
        company = ''
        client_name = ''
        phone = ''
        email = ''

    return {
        'Company': company,
        'Client Name': client_name,
        'Phone': phone,
        'Email': email
    }

# Function to extract invoice data
def extract_invoice_data(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text()
    
    # Debug print the full text
    print("Full text extracted from PDF:")
    print(text)
    
    # Extract Bill To information
    bill_to_info = extract_bill_to_info(text)
    
    # Extract Invoice #
    invoice_number = extract_invoice_number(pdf_path)
    
    # Determine Delivery Method
    if "Delivery" in text:
        delivery_method = "海外仓"
    else:
        delivery_method = "自提"
    
    return {
        'Invoice #': invoice_number,
        'Bill To': bill_to_info,
        'Delivery Method': delivery_method
    }

# Function to extract tables from a specified area in the PDF
def extract_table_from_bbox(pdf_path, bbox):
    unique_rows = set()  # Set to store all unique rows
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Crop the page to the bounding box area
            cropped_page = page.within_bbox(bbox)
            tables = cropped_page.extract_tables()
            for table in tables:
                cleaned_table = clean_table(table)
                formatted_rows = format_table(cleaned_table)
                unique_rows.update(formatted_rows)  # Add formatted rows to the unique set
    
    # Filter rows starting with a number
    filtered_rows = [row for row in unique_rows if re.match(r'^\d', row)]
    
    # Process rows into DataFrame format
    records = []
    for row in filtered_rows:
        qty_match = re.match(r'^(\d+)', row)
        if qty_match:
            qty = qty_match.group(1)
            description_start = row.find('- Description:') + len('- Description:')
            item_description = row[qty_match.end():description_start - len('- Description:')].strip()
            description = row[description_start:].strip()
            
            # Exclude items starting with 'assemble' or 'deliver'
            if not item_description.lower().startswith(('assem', 'deliver')):
                records.append({'Qty': qty, 'Item': item_description, 'Description': description})
    
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Sort by 'Item' column in ascending order
    df.sort_values(by='Item', inplace=True)
    
    return df

# Function to clean table rows
def clean_table(table):
    cleaned_table = []
    for row in table:
        if len(row) > 1:  # Ensure the row has more than one element (to skip headers)
            cleaned_row = [cell.strip() if cell else '' for cell in row]  # Strip whitespace
            cleaned_table.append(cleaned_row)
    return cleaned_table

# Function to format table rows
def format_table(cleaned_table):
    formatted_rows = set()  # Use a set to avoid duplicates
    for row in cleaned_table:
        if len(row) >= 2:
            item_description = ' '.join(row[:-1])  # Combine all columns except the last one
            qty = row[-1]  # The last column is the quantity
            formatted_row = f"{item_description} - Description: {qty}"
            formatted_rows.add(formatted_row)  # Add to set to ensure uniqueness
    return formatted_rows

# Function to combine invoice data and table data into a single structure
def combine_data(invoice_data, products_df):
    # Modify item names based on description and remove special characters
    for idx, row in products_df.iterrows():
        item_name = row['Item'].replace(" ", "-")
        item_name = re.sub(r'[^A-Za-z0-9-]', '', item_name)  # Remove special characters except dashes
        
        if row['Description'].startswith(('Ace', 'Acce','Touch up')):
            products_df.at[idx, 'Item'] = f"Accessory-{item_name}"
        elif row['Description'].startswith(('Stone', 'Premium Stone')):
            products_df.at[idx, 'Item'] = f"Stone-{item_name}"
        else:
            products_df.at[idx, 'Item'] = f"Cabinet-{item_name}"

    data = {
        'Invoice #': invoice_data['Invoice #'],
        'Company': invoice_data['Bill To']['Company'],
        'Client Name': invoice_data['Bill To']['Client Name'],
        'Phone': invoice_data['Bill To']['Phone'],
        '*Warehouse': 'USTEST',
        '*Order Type': 'PO订单',
        '*Delivery Method': invoice_data['Delivery Method'],
        '*Logistic Type': 'Truck',
        '*Packing Type': '仓库打包',
        'SKU and QTY': products_df[['Item', 'Qty']].to_dict('records')
    }
    return data

# Function to create the Excel template
def create_template(path):
    wb = Workbook()
    ws = wb.active

    # Define column names for order information
    order_columns = [
        '*Order number', '*Warehouse', 'Store Name', '*Order Type', '*Delivery Method', '*Logistic Type', '*Packing Type',
        'FBA Code', 'Company', 'Consignee', 'Country', 'State', 'Phone', 'City', 'Zip', 'Street1', 'Street2', 'Pickup At',
        'Pickup Number', 'Pallet Qty.'
    ]

    # Define column names for product details
    product_columns = [
        '*SKU', '*Total Qty.', 'Quantity In Box', 'Box Qty', '*Change sku label', 'New SKU', 'Box reference'
    ]

    # Set the font to bold
    bold_font = Font(bold=True)

    # Write the order columns (merged A1:T2) with bold font and adjust column widths
    for col_num, column_title in enumerate(order_columns, 1):
        ws.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        ws.column_dimensions[cell.column_letter].width = len(column_title) + 2  # Adjust width based on header length

    # Write the product list title (merged U1:AA1) with bold font and adjust column width
    ws.merge_cells(start_row=1, start_column=21, end_row=1, end_column=27)
    cell = ws.cell(row=1, column=21, value='z.transfer.skuList')
    cell.font = bold_font
    ws.column_dimensions[cell.column_letter].width = len('z.transfer.skuList') + 2  # Adjust width based on header length

    # Write the product columns (U2:AA2) with bold font and adjust column widths
    for col_num, column_title in enumerate(product_columns, 21):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = bold_font
        ws.column_dimensions[cell.column_letter].width = len(column_title) + 2  # Adjust width based on header length

    # Write the description title (merged AB1:AB2) with bold font and adjust column width
    ws.merge_cells(start_row=1, start_column=28, end_row=2, end_column=28)
    cell = ws.cell(row=1, column=28, value='Description')
    cell.font = bold_font
    ws.column_dimensions[cell.column_letter].width = len('Description') + 2  # Adjust width based on header length

    wb.save(path)

# Function to map data to the Excel template
def map_data_to_template(invoice_data, products_df, ws, start_row):
    # Write order information
    ws.cell(row=start_row, column=1, value=invoice_data['Invoice #'])
    ws.cell(row=start_row, column=2, value=invoice_data['*Warehouse'])
    ws.cell(row=start_row, column=3, value="")
    ws.cell(row=start_row, column=4, value=invoice_data['*Order Type'])
    ws.cell(row=start_row, column=5, value=invoice_data['*Delivery Method'])
    ws.cell(row=start_row, column=6, value=invoice_data['*Logistic Type'])
    ws.cell(row=start_row, column=7, value=invoice_data['*Packing Type'])
    ws.cell(row=start_row, column=8, value="")
    ws.cell(row=start_row, column=9, value=invoice_data['Company'])
    ws.cell(row=start_row, column=10, value=invoice_data['Client Name'])
    ws.cell(row=start_row, column=11, value="")
    ws.cell(row=start_row, column=12, value="")
    ws.cell(row=start_row, column=13, value=invoice_data['Phone'])
    ws.cell(row=start_row, column=14, value="")
    ws.cell(row=start_row, column=15, value="")
    ws.cell(row=start_row, column=16, value="")
    ws.cell(row=start_row, column=17, value="")
    ws.cell(row=start_row, column=18, value="")
    ws.cell(row=start_row, column=19, value="")
    ws.cell(row=start_row, column=20, value="")

    # Map products to the template
    for i, row in products_df.iterrows():
        ws.cell(row=start_row + i, column=21, value=row['Item'])
        ws.cell(row=start_row + i, column=22, value=row['Qty'])
        ws.cell(row=start_row + i, column=28, value=row['Description'])

    return start_row + len(products_df)  # Return the next starting row without adding an empty row

# Main function to process PDFs and generate the Excel template
def main():
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Make the root window appear in front of all other windows
    root.update()
    pdf_paths = filedialog.askopenfilenames(title="Select Invoice PDF files", filetypes=[("PDF Files", "*.pdf")])
    if not pdf_paths:
        messagebox.showwarning("No file selected", "No PDF file selected. Exiting...")
        return

    template_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Forwarding Import Template", filetypes=[("Excel files", "*.xlsx")])
    if not template_path:
        messagebox.showwarning("No save location", "No save location specified. Exiting...")
        return

    root.attributes('-topmost', False)  # Reset the root window attribute
    root.after(500, root.destroy)  # Delay the destruction of the root window to allow dialogs to complete

    # Create the template
    create_template(template_path)
    wb = load_workbook(template_path)
    ws = wb.active

    start_row = 3  # Starting row for the first order
    for pdf_path in pdf_paths:
        invoice_data = extract_invoice_data(pdf_path)
        bbox = (0, 280, 612, 792)  # Adjust the bounding box values as needed
        products_df = extract_table_from_bbox(pdf_path, bbox)
        combined_data = combine_data(invoice_data, products_df)
        start_row = map_data_to_template(combined_data, products_df, ws, start_row)

    wb.save(template_path)
    print(f"Template created and saved as {template_path}")
    messagebox.showinfo("Success", f"Template created and saved as {template_path}")

# Run the main function
if __name__ == "__main__":
    main()
